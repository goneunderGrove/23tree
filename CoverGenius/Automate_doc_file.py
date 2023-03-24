
#    /$$$$$$                                           /$$$$$$                      /$$
#   /$$__  $$                                         /$$__  $$                    |__/
#  | $$  \__/  /$$$$$$  /$$    /$$ /$$$$$$   /$$$$$$ | $$  \__/  /$$$$$$  /$$$$$$$  /$$ /$$   /$$  /$$$$$$$
#  | $$       /$$__  $$|  $$  /$$//$$__  $$ /$$__  $$| $$ /$$$$ /$$__  $$| $$__  $$| $$| $$  | $$ /$$_____/
#  | $$      | $$  \ $$ \  $$/$$/| $$$$$$$$| $$  \__/| $$|_  $$| $$$$$$$$| $$  \ $$| $$| $$  | $$|  $$$$$$
#  | $$    $$| $$  | $$  \  $$$/ | $$_____/| $$      | $$  \ $$| $$_____/| $$  | $$| $$| $$  | $$ \____  $$
#  |  $$$$$$/|  $$$$$$/   \  $/  |  $$$$$$$| $$      |  $$$$$$/|  $$$$$$$| $$  | $$| $$|  $$$$$$/ /$$$$$$$/
#   \______/  \______/     \_/    \_______/|__/       \______/  \_______/|__/  |__/|__/ \______/ |_______/
#
#
#



"""
This code reads data from an Excel workbook and uses it to fill in placeholders in a Word document to create a cover letter.
It then saves the resulting cover letter and a general resume in a directory named after the company being applied to.
"""

import os
import shutil

from openpyxl import load_workbook
from docx import Document
from docxtpl import DocxTemplate

print("""
#   _____
#  /  __ \
#
#  | /  \/ _____   _____ _ __
#  | |    / _ \ \ / / _ \ '__|
#  | \__/\ (_) \ V /  __/ |
#   \____/\___/ \_/ \___|_|
#  _____            _
# |  __ \          (_)
# | |  \/ ___ _ __  _ _   _ ___
# | | __ / _ \ '_ \| | | | / __|
# | |_\ \  __/ | | | | |_| \__ \
#
#  \____/\___|_| |_|_|\__,_|___/
#
#
""")


# Define file paths
main_path = r"../resume"
template_path = os.path.join(main_path, 'coverlettertemp.docx')
workbook_path = os.path.join(main_path, 'Job_postings.xlsx')

# Load files
workbook = load_workbook(workbook_path)
template = DocxTemplate(template_path)

print(workbook.sheetnames)

#Iterate through each sheet in the workbook
for worksheet in workbook.sheetnames:
	worksheet = workbook['Sheet1']
	print(worksheet)

# Define placeholders to be filled in the Word document
to_fill_in = {
               'Hiring_Manager_First_Name':'Hiring_Manager_First_Name',
               'Hiring_Manager_Last_Name' :'Hiring_Manager_Last_Name',
               'Applicant_First_Name':'Applicant_First_Name',
               'Applicant_Last_Name':'Applicant_Last_Name',
               'Website':'Website',
               'Company':'Company',
               'Position':'Position',
               'year_of_experience':'year_of_experience',
               'specific_skills_or_technologies':'specific_skills_or_technologies',
               'specific_projects_or_responsibilities':'specific_projects_or_responsibilities',
               'interest_n_company': 'interest_n_company'
              }

# Set the minimum number of columns. This will be 2.
column = 2

# print out the maximum columns that are filled in in the excel file. This is to see how many iterations the code will need.
print(worksheet.max_column)

# Perform the following code block if the colomn amoumnt is less than the maximum column amount.
while column <= worksheet.max_column:

   # Define the column index. This is a letter so you need to convert the column number to a letter (2+64) = B
   col_index = chr(column+64)
   row_index = 1
   # Retrieve the values from excel document and store in dictionary you defined earlier on
   # For each key in the dictionary, we look up the value in the excel file and store it instead of "none" in the dictionary
   for key in to_fill_in:
       cell = '%s%i' % (col_index, row_index)
       to_fill_in[key] = worksheet[cell].value
       row_index += 1
    
   # Fill in all the keys defined in the word document using the dictionary.
   # The keys in de word document are identified by the {{}}symbols.
   template.render(to_fill_in)
   
   # Output the file to a docx document.
   directory_path = "../resume/Finished_CoverLetter/"+ str(to_fill_in['Company'])
   if not os.path.exists(directory_path):
      os.makedirs(directory_path)
      filename = str(to_fill_in['Company']) + '_Cover_letter.docx'
      save_path = r"../resume/Finished_CoverLetter/"+ str(to_fill_in['Company'])
      filled_path = os.path.join(save_path, filename)
      template.save(filled_path)
      src_file = "../resume/general_resume.docx"
      dst_dir = save_path
      new_file_name = str(to_fill_in['Company']) + '_Resume.docx'
      shutil.copy(src_file, dst_dir+"/"+new_file_name)
   else:
    print(f"{directory_path} already exists.")
   print("Done with %s" % str(to_fill_in['Company']))
   column +=1
   

